# -*- coding:utf-8 -*-
'''
简化用户数据记录
'''
from datetime import date
import openpyxl
import os


file_path = "/Users/wangzai/PycharmProjects/ZaiFirst/TGM/User.xlsx"
"用户资产表地址"
workbook = openpyxl.load_workbook(file_path)
"加载表格"
sheet1 = workbook.worksheets[0]
"加载sheet"
row = sheet1.max_row
"获取行数"
column = sheet1.max_column
"获取列"
row_value = []
"行内容"
name = ["雪舟", "小宁", "女侠", "ruby", "邓哥", "乐绡", "学姐", "丘", "小米", "子颖", "Ma", "深深", "wangqian", "千玥", "Muzi"]
"参加活动的用户名"
locat = "奥森"
# date = date.today()
date = "2023-06-8"
"参见活动的时间"

for j in range(0,row):
    cell_value = sheet1.cell(row=j+1,column=1).value
    row_value.append(cell_value)
"获取所有的用户名"
for i in range(0,len(name)):
    "根据用户名判断是否在表中"
    if name[i] in row_value:
        x = row_value.index(name[i])
        value = sheet1.cell(x,2).value
        sheet1.cell(x,2,value + 1)
        "活动次数 + 1"
        sheet1.cell(x,3,date)
        "更新活动时间"
    else:
        sheet1.append([name[i],1,date,3,3,3,3,3,"mid","deep","男",locat,date,1])
        "不在表中添加新的会员数据"
workbook.save("User.xlsx")
workbook.close()
# print(row_value)